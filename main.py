import requests
import re
import customtkinter as ctk
from tkinter import filedialog, messagebox
import tkinter as tk
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def on_futtatas():
    try:
        adatok = get_input_xlsx(path_var.get())
        osszes_sor = len(adatok)

        final_data = [("Csomagszám", "Irányítószám", "Indulás dátuma", "Kézbesítés dátuma")]

        # --- GUI Setup ---
        progress_window = ctk.CTkToplevel(root)
        progress_window.title("Feldolgozás...")
        progress_window.geometry("300x150")

        x = root.winfo_x() + (root.winfo_width() // 2) - 150
        y = root.winfo_y() + (root.winfo_height() // 2) - 75
        progress_window.geometry(f"+{x}+{y}")
        progress_window.transient(root)
        progress_window.grab_set()

        lbl_status = ctk.CTkLabel(progress_window, text=f"Indítás: {osszes_sor} sor...", font=("Roboto", 13))
        lbl_status.pack(pady=(25, 10))

        progress_bar = ctk.CTkProgressBar(progress_window, width=220)
        progress_bar.set(0)
        progress_bar.pack(pady=5)
        root.update()
        
        completed_count = 0

        for csomagszam, iranyitoszam in adatok:
            if csomagszam.startswith("KN-"):
                departure_date, arrival_date = get_kn_status(csomagszam.upper().strip("KN-"))
            else:
                departure_date, arrival_date = get_gls_status(csomagszam.upper().strip("GLS-"), iranyitoszam)

            final_data.append((csomagszam, iranyitoszam, departure_date, arrival_date))
            print(f"Csomagszám: {csomagszam}, Irányítószám: {iranyitoszam}, Státusz: {departure_date}, {arrival_date}")

            completed_count += 1
            szazalek = completed_count / osszes_sor
            progress_bar.set(szazalek)
            lbl_status.configure(text=f"Kész: {completed_count} / {osszes_sor}")
            root.update()

    except Exception as e:
        hiba_ablak(f"Kritikus hiba történt: {e}")

    finally:
        lbl_status.configure(text="Mentés fájlba...")
        root.update()

        wb = Workbook()
        ws = wb.active
        for sor in final_data:
            ws.append(sor)
        
        # Formázás
        for cell in ws[1]:
            cell.font = Font(bold=True, size=12)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.freeze_panes = "A2"
        wb.save(f'{mezo.get()}.xlsx')

        progress_window.grab_release()
        progress_window.destroy()
        root.destroy()

def get_input_xlsx(file):
    adatok = []
    
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        sheet = wb.active
        
        for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
            csomagszam, iranyitoszam = row
            
            if csomagszam is not None:
                adatok.append((str(csomagszam).strip(), str(iranyitoszam).strip()))
                
        return adatok

    except Exception as e:
        print(f"Váratlan hiba történt a fájl olvasásakor: {e}")
        return []

def get_kn_status(tracking_number):
    session = requests.Session()
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8"
    }

    try:
        page_url = f"https://mykn.kuehne-nagel.com/public-tracking/shipments?query={tracking_number}" 
        
        response = session.get(page_url, headers=headers)
        response.raise_for_status()
        
        html_content = response.text
        match = re.search(r'/shipments/(\d+)', html_content)
        
        if not match:
            return "Nem található meg az internal_id a HTML-ben! Lehet, hogy a struktúra megváltozott vagy a szám nem helyes."
            
        internal_id = match.group(1)

        routing_url = f"https://mykn.kuehne-nagel.com/public-tracking/internal/shipments/{internal_id}/shipment-routing"

        api_response = session.get(routing_url, headers=headers)
        api_response.raise_for_status()
        
        routing_data = api_response.json()

        arrival_date = "Még nem érkezett meg."
    
        if routing_data["routeLocations"][-1]["completed"] == True:
            arrival_date = routing_data["routeLocations"][-1]["locationMilestones"][-1]["actualAchievementDateTime"]["dateTime"]["date"]

        departure_date = routing_data["routeLocations"][0]["locationMilestones"][0]["actualAchievementDateTime"]["dateTime"]["date"]
        
        # print(f"Bumm, megvan a dátum! Az érték: {departure_date}, {arrival_date}")
        return departure_date, arrival_date

    except requests.exceptions.RequestException as e:
        return f"Hálózati hiba: {e}"

def get_gls_status(tracking_number, postal_code):
    url = f"https://gls-group.com/app/service/open/rest/HU/hu/rstt028/{tracking_number}?caller=witt002&millis=1772520728102&tuOwnerCode=HU01&postalCode={postal_code}"
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        
        data = response.json()
        
        arrival_date = "Még nem érkezett meg."

        if data["history"][0]["evtDscr"] == "A csomag kézbesítésre ker&#252;lt.":
            arrival_date = data["history"][0]["date"]

        departure_date = data["history"][-2]["date"]

        # print(f"Bumm, megvan a dátum! Az érték: {departure_date}, {arrival_date}")
        return departure_date, arrival_date

    except requests.exceptions.RequestException as e:
        print(f"Hiba történt a kérés során: {e}")
        return ("Hiba", "Hálózati Hiba vagy hibás adatok")

def hiba_ablak(uzenet):
    root = tk.Tk(); root.withdraw(); messagebox.showerror("Hiba történt!", uzenet); root.destroy()

ctk.set_appearance_mode("Dark")
ctk.set_default_color_theme("dark-blue")

if __name__ == "__main__":

    root = ctk.CTk()
    root.geometry("360x320") 
    root.title("Csomagkövető")
    root.resizable(False, False)

    my_font_head = ("Roboto", 14)
    my_font_btn = ("Roboto", 13)
    grey_btn_color = "gray30" 
    grey_hover_color = "gray40"

    path_var = ctk.StringVar()
    filename_var = ctk.StringVar(value="output")

    def gomb_allapot_ellenorzes(*args):
        mappa_van = len(path_var.get()) > 0
        fajlnev_van = len(filename_var.get()) > 0

        if mappa_van and fajlnev_van:
            btn_run.configure(state="normal", fg_color=grey_btn_color)
        else:
            btn_run.configure(state="disabled", fg_color="gray20")
    

    input_frame = ctk.CTkFrame(root, fg_color="transparent")
    input_frame.pack(pady=(0, 2), fill="x")

    ctk.CTkLabel(input_frame, text="Bemeneti fájl:", font=my_font_head, text_color="gray80").pack(pady=(20, 5))
    btn_browse = ctk.CTkButton(input_frame, text="Excel fájl választás",
                               command=lambda: path_var.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])),
                               font=my_font_btn,
                               fg_color=grey_btn_color, hover_color=grey_hover_color)
    btn_browse.pack(pady=5)

    path_label = ctk.CTkLabel(input_frame, textvariable=path_var, font=("Arial", 11), text_color="gray60", wraplength=340)

    def on_path_change(*args):
        if path_var.get():
             path_label.pack(padx=10, pady=(0))
        else:
             path_label.pack_forget()
        gomb_allapot_ellenorzes()

    path_var.trace_add("write", on_path_change)


    ctk.CTkLabel(root, text="Kimeneti fájl neve:", font=my_font_head, text_color="gray80").pack(pady=(15, 5))
    mezo = ctk.CTkEntry(root, width=200, justify='center', font=my_font_btn, textvariable=filename_var)
    filename_var.trace_add("write", gomb_allapot_ellenorzes)
    mezo.pack(padx=10, pady=5)

    btn_run = ctk.CTkButton(root, text="Futtatás", command=on_futtatas,
                            font=("Roboto", 16, "bold"), height=40,
                            state="disabled",
                            fg_color="gray20",
                            hover_color=grey_hover_color)
    btn_run.pack(pady=(30, 20))

    gomb_allapot_ellenorzes()
    root.mainloop()